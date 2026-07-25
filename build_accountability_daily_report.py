#!/usr/bin/env python3
"""
Rebuilds the JAIN-E Accountability Daily Usage Report (Active Only) workbook
from scratch, in the exact visual format of the reference template, and
OVERWRITES the given output path (no new dated file is ever created).

Usage:
    python3 build_accountability_daily_report.py <input_json> <output_xlsx> "<date_label>" <active_count_yesterday_placeholder_unused>

<input_json> must be a JSON array of objects, one per ACTIVE user (i.e. only
users who created / received / approved a task or added a comment on the
target day), each with keys:
    full_name          str
    email               str
    department          str   (already comma-joined if multiple)
    created_titles      str or null   (" | "-joined task titles)
    received_titles     str or null   (" | "-joined "title (from email)")
    approved_titles     str or null   (" | "-joined task titles)
    comment_snippets    str or null   (" | "-joined snippets, already truncated)

<date_label> e.g. "Wednesday, 22 July 2026"
"""
import sys, json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def main():
    in_json, out_path, date_label = sys.argv[1], sys.argv[2], sys.argv[3]
    with open(in_json, encoding="utf-8") as f:
        rows = json.load(f)
    rows = sorted(rows, key=lambda r: (r.get("full_name") or r.get("email") or "").lower())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Accountability Usage"

    FONT = "Arial"
    NAVY = "0A2640"
    SLATE = "475569"
    GREEN_FILL, GREEN_TXT = "DCFCE7", "166534"
    RED_FILL, RED_TXT = "FEE2E2", "991B1B"
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["#","Full Name","Email","Department","Created Task","Created — Task Titles",
               "Received Task","Received — Task Titles","Approved Task","Approved — Task Titles",
               "Comment Added — Snippets"]

    ws.merge_cells("A1:K1")
    c = ws["A1"]; c.value = "JAIN-E Accountability — Daily Usage Report"
    c.font = Font(name=FONT, size=14, bold=True, color=NAVY)
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:K2")
    c = ws["A2"]; c.value = f"Active users only — {date_label} (IST) — Accountability module only"
    c.font = Font(name=FONT, size=10, color=SLATE)
    ws.row_dimensions[2].height = 16

    HEADER_ROW = 4
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=HEADER_ROW, column=i, value=h)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[HEADER_ROW].height = 30

    start_row = HEADER_ROW + 1
    for idx, r in enumerate(rows):
        row = start_row + idx
        created = bool(r.get("created_titles"))
        received = bool(r.get("received_titles"))
        approved = bool(r.get("approved_titles"))
        comments = r.get("comment_snippets") or "—"

        ws.cell(row=row, column=1, value=idx+1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=r.get("full_name") or "")
        ws.cell(row=row, column=3, value=r.get("email") or "")
        ws.cell(row=row, column=4, value=r.get("department") or "")

        def yn_cell(col, flag):
            cell = ws.cell(row=row, column=col, value="Yes" if flag else "No")
            cell.font = Font(name=FONT, size=10.5, bold=True, color=GREEN_TXT if flag else RED_TXT)
            cell.fill = PatternFill("solid", fgColor=GREEN_FILL if flag else RED_FILL)
            cell.alignment = Alignment(horizontal="center")

        yn_cell(5, created)
        ws.cell(row=row, column=6, value=r.get("created_titles") or "")
        yn_cell(7, received)
        ws.cell(row=row, column=8, value=r.get("received_titles") or "")
        yn_cell(9, approved)
        ws.cell(row=row, column=10, value=r.get("approved_titles") or "")
        ws.cell(row=row, column=11, value=comments)

        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if cell.font.name is None:
                cell.font = Font(name=FONT, size=10.5)
            if col in (2,3,4):
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col in (6,8,11):
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    last_row = start_row + len(rows) - 1 if rows else start_row - 1

    widths = {1:5, 2:20, 3:30, 4:14, 5:12, 6:34, 7:12, 8:34, 9:12, 10:34, 11:34}
    for col, w in widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    ws.freeze_panes = f"A{start_row}"

    sum_row1 = last_row + 2
    sum_row2 = sum_row1 + 1
    note_row = sum_row2 + 1

    c = ws.cell(row=sum_row1, column=1, value="Summary")
    c.font = Font(name=FONT, size=11, bold=True, color=NAVY)

    c = ws.cell(row=sum_row2, column=1, value="Active users yesterday:")
    c.font = Font(name=FONT, size=10, bold=True)
    cval = ws.cell(row=sum_row2, column=2,
                    value=(f"=COUNTA(A{start_row}:A{last_row})" if rows else 0))
    cval.font = Font(name=FONT, size=10, bold=False)

    ws.merge_cells(f"A{note_row}:K{note_row}")
    c = ws.cell(row=note_row, column=1,
                value='Note: "Created" and "Received" are mutually exclusive per task — a task you '
                      'created and self-assigned counts only under Created, not Received. "Received" '
                      'only counts tasks someone else delegated to you.')
    c.font = Font(name=FONT, size=9, color=SLATE)
    c.alignment = Alignment(wrap_text=True)

    wb.save(out_path)
    print("saved:", out_path, "rows:", len(rows))

if __name__ == "__main__":
    main()
